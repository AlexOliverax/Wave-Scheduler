import os
import pandas as pd
import numpy as np
import logging
from datetime import datetime, timedelta
import re
from collections import defaultdict
from app.utils import get_country_holidays_dict

logger = logging.getLogger(__name__)

class WavesScheduler:
    """
    Core class for the Waves Scheduler application.
    """
    
    def __init__(self):
        """Initialize the scheduler."""
        self.data = None
        self.total_devices = 0
    
    def load_file(self, file_path, preview_only=False, selected_columns=None):
        """
        Load data from a CSV or Excel file.
        
        Args:
            file_path (str): Path to the data file
            preview_only (bool): If True, only load the file to get columns
            selected_columns (list): List of columns to include
            
        Returns:
            bool: True if file was loaded successfully, False otherwise
        """
        try:
            # Determine file type based on extension
            file_ext = os.path.splitext(file_path)[1].lower()
            
            if file_ext == '.csv':
                # Try different encodings
                try:
                    self.data = pd.read_csv(file_path, encoding='utf-8')
                except UnicodeDecodeError:
                    self.data = pd.read_csv(file_path, encoding='latin1')
            elif file_ext in ['.xlsx', '.xls', '.xlsm']:
                self.data = pd.read_excel(file_path)
            else:
                logger.error(f"Unsupported file format: {file_ext}")
                return False
            
            # If preview only, return now (no column validation)
            if preview_only:
                return True
            
            # Filter columns if specified
            if selected_columns:
                self.data = self.data[selected_columns]
            
            # Count total devices
            self.total_devices = len(self.data)
            logger.info(f"Loaded {self.total_devices} devices from {file_path}")
            
            return True
            
        except Exception as e:
            logger.error(f"Error loading file: {str(e)}")
            return False
    
    def calculate_devices_per_wave(self, bandwidth_mb, mb_per_device=0.5):
        """
        Calculate the number of devices per wave based on bandwidth.
        
        Args:
            bandwidth_mb (float): Available bandwidth in MB
            mb_per_device (float): MB required per device (default: 0.5)
            
        Returns:
            int: Number of devices per wave
        """
        # Calculate devices per wave based on configurable MB per device
        devices_per_wave = int(bandwidth_mb / mb_per_device)
        
        # Ensure at least 1 device per wave
        return max(1, devices_per_wave)
    
    def calculate_ideal_waves(self, total_devices, devices_per_wave):
        """
        Calculate the ideal number of waves.
        
        Args:
            total_devices (int): Total number of devices
            devices_per_wave (int): Number of devices per wave
            
        Returns:
            int: Ideal number of waves
        """
        return max(1, int(np.ceil(total_devices / devices_per_wave)))
    
    def generate_wave_labels(self, start_date, num_waves, avoid_holidays=True, country_code="BR"):
        """
        Gera rótulos de waves baseados na data inicial, pulando finais de semana e feriados.

        Args:
            start_date (datetime): Data inicial para a primeira wave
            num_waves (int): Número de waves
            avoid_holidays (bool): Se True, pula feriados nacionais do país configurado
            country_code (str): Código do país para verificação de feriados

        Returns:
            list: Lista de rótulos de waves
        """
        wave_labels = []
        current_date = start_date
        wave_count = 0

        # Pré-carregar feriados do ano (e do próximo, caso ultrapasse)
        holidays_set = set()
        if avoid_holidays:
            for yr in {start_date.year, start_date.year + 1}:
                holidays_set.update(get_country_holidays_dict(yr, country_code).keys())

        while wave_count < num_waves:
            current_day = current_date.date() if hasattr(current_date, 'date') else current_date
            weekday = current_date.weekday()
            is_weekend = weekday >= 5
            is_holiday = avoid_holidays and current_day in holidays_set

            if not is_weekend and not is_holiday:
                wave_label = f"Wave {wave_count+1} - {current_date.strftime('%d/%m/%Y')}"
                wave_labels.append(wave_label)
                wave_count += 1

            current_date += timedelta(days=1)

        return wave_labels
    
    def _group_similar_devices(self):
        """
        Group devices by similarity in path or device name.
        
        Returns:
            dict: Dictionary of device groups
        """
        groups = defaultdict(list)
        
        # Try to find columns for grouping (flexible column names)
        path_col = None
        device_col = None
        
        # Look for path-like columns
        for col in self.data.columns:
            if 'path' in col.lower():
                path_col = col
                break
        
        # Look for device name-like columns
        for col in self.data.columns:
            if 'device' in col.lower() or 'name' in col.lower():
                device_col = col
                break
        
        # Extract location from path (usually first part of path)
        for idx, row in self.data.iterrows():
            location = "default"
            prefix = ""
            
            # Try to extract location from path column
            if path_col and path_col in row:
                path = str(row[path_col])
                # Extract location from path (first part of path)
                location = path.split('\\')[0] if '\\' in path else path.split('/')[0]
            
            # Try to extract prefix from device name
            if device_col and device_col in row:
                device_name = str(row[device_col])
                # Extract prefix from device name (e.g., "DESKTOP-" or "LAPTOP-")
                prefix_match = re.match(r'^([A-Za-z]+-)', device_name)
                prefix = prefix_match.group(1) if prefix_match else ""
            
            # Create a group key based on location and prefix
            group_key = f"{location}_{prefix}"
            groups[group_key].append(idx)
        
        return groups
    
    def distribute_devices(self, num_waves, devices_per_wave):
        """
        Distribute devices into waves, ensuring similar devices are spread across waves.
        
        Args:
            num_waves (int): Number of waves
            devices_per_wave (int): Maximum number of devices per wave
            
        Returns:
            dict: Dictionary mapping wave labels to lists of device indices
        """
        if self.data is None or self.total_devices == 0:
            logger.error("No data loaded")
            return {}
        
        # Group similar devices
        device_groups = self._group_similar_devices()
        
        # Initialize waves
        waves = [[] for _ in range(num_waves)]
        wave_counts = [0] * num_waves
        
        # Distribute groups across waves
        for group_key, device_indices in sorted(device_groups.items(), key=lambda x: len(x[1]), reverse=True):
            # For each device in the group
            for idx in device_indices:
                # Find the wave with the fewest devices
                target_wave = wave_counts.index(min(wave_counts))
                
                # Add device to wave if not full
                if wave_counts[target_wave] < devices_per_wave:
                    waves[target_wave].append(idx)
                    wave_counts[target_wave] += 1
                else:
                    # Find next available wave
                    for w in range(num_waves):
                        if wave_counts[w] < devices_per_wave:
                            waves[w].append(idx)
                            wave_counts[w] += 1
                            break
        
        # Create wave distribution dictionary
        wave_distribution = {}
        for i, wave_devices in enumerate(waves):
            if wave_devices:  # Only include waves with devices
                wave_label = f"Wave {i+1}"
                wave_distribution[wave_label] = [self.data.iloc[idx].to_dict() for idx in wave_devices]
        
        logger.info(f"Distributed {self.total_devices} devices into {len(wave_distribution)} waves")
        return wave_distribution
    
    def generate_csv(self, output_path, wave_distribution, wave_labels):
        """
        Generate a CSV file with the wave distribution.
        Each device row gets a 'Wave' column indicating its assigned wave.

        Args:
            output_path (str): Path to save the CSV file
            wave_distribution (dict): Dictionary mapping wave labels to lists of device dicts
            wave_labels (list): List of wave labels with dates

        Returns:
            bool: True if CSV file was generated successfully, False otherwise
        """
        try:
            rows = []
            for i, wave_label in enumerate(wave_labels):
                wave_key = f"Wave {i+1}"
                if wave_key in wave_distribution:
                    for device in wave_distribution[wave_key]:
                        row = {"Wave": wave_label}
                        row.update(device)
                        rows.append(row)

            if not rows:
                logger.warning("No rows to export to CSV")
                return False

            import pandas as pd
            df = pd.DataFrame(rows)
            df.to_csv(output_path, index=False, encoding="utf-8-sig")
            logger.info(f"CSV file generated: {output_path}")
            return True
        except Exception as e:
            logger.error(f"Error generating CSV file: {str(e)}")
            return False

    def generate_excel(self, output_path, wave_distribution, wave_labels, rfc=None):
        """
        Generate an Excel file with the wave distribution.
        
        Args:
            output_path (str): Path to save the Excel file
            wave_distribution (dict): Dictionary mapping wave labels to lists of device indices
            wave_labels (list): List of wave labels
            rfc (str, optional): RFC number
            
        Returns:
            bool: True if Excel file was generated successfully, False otherwise
        """
        try:
            # Create Excel writer
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Create summary sheet
                summary_data = {
                    'Wave': [],
                    'Date': [],
                    'Devices': [],
                    'RFC': []
                }
                
                for i, wave_label in enumerate(wave_labels):
                    wave_key = f"Wave {i+1}"
                    if wave_key in wave_distribution:
                        devices = wave_distribution[wave_key]
                        date_part = wave_label.split(' - ')[1]
                        # Substituir barras por hífens na data
                        date_part = date_part.replace('/', '-')
                        summary_data['Wave'].append(f"Wave {i+1}")
                        summary_data['Date'].append(date_part)
                        summary_data['Devices'].append(len(devices))
                        summary_data['RFC'].append(rfc if rfc else "N/A")
                
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
                
                # Formatar o cabeçalho da planilha Summary
                workbook = writer.book
                summary_sheet = writer.sheets['Summary']
                self._format_header(summary_sheet)
                
                # Create sheet for each wave
                for i, wave_label in enumerate(wave_labels):
                    wave_key = f"Wave {i+1}"
                    if wave_key in wave_distribution:
                        devices = wave_distribution[wave_key]
                        date_part = wave_label.split(' - ')[1]
                        # Substituir barras por hífens na data
                        date_part = date_part.replace('/', '-')
                        # Usar o nome da aba com a data
                        sheet_name = f"Wave {i+1} {date_part}"
                        # Limitar o tamanho do nome da aba para 31 caracteres (limite do Excel)
                        if len(sheet_name) > 31:
                            sheet_name = sheet_name[:31]
                        wave_df = pd.DataFrame(devices)
                        wave_df.to_excel(writer, sheet_name=sheet_name, index=False)
                        
                        # Formatar o cabeçalho da planilha da wave
                        wave_sheet = writer.sheets[sheet_name]
                        self._format_header(wave_sheet)
            
            logger.info(f"Excel file generated: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error generating Excel file: {str(e)}")
            return False
    
    def _format_header(self, worksheet):
        """
        Format the header row of a worksheet with professional styling.

        Args:
            worksheet: The openpyxl worksheet object
        """
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        # Cores premium
        header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
        white_font = Font(color='FFFFFF', bold=True, name='Segoe UI', size=10)
        centered = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='FFFFFF'),
            right=Side(style='thin', color='FFFFFF'),
            bottom=Side(style='medium', color='4F76FF'),
        )

        # Formatação do cabeçalho
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = centered
            cell.border = thin_border

        # Altura da linha do cabeçalho
        worksheet.row_dimensions[1].height = 22

        # Zebra striping nas linhas de dados
        light_fill = PatternFill(start_color='F0F4FF', end_color='F0F4FF', fill_type='solid')
        data_font = Font(name='Segoe UI', size=9)
        data_alignment = Alignment(vertical='center')
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
            fill = light_fill if row_idx % 2 == 0 else None
            for cell in row:
                if fill:
                    cell.fill = fill
                cell.font = data_font
                cell.alignment = data_alignment

        # Ajustar a largura das colunas automaticamente
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            worksheet.column_dimensions[column_letter].width = min(max_length + 4, 50)